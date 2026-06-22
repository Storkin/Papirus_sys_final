from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.contrib.auth.views import LoginView
from django.contrib.auth.decorators import login_required
from .models import Product, StockMovement, Customer, Sale, SaleItem, ReferenceBarcode
from .forms import ProductForm, CustomerForm
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.urls import reverse
from django.db.models import Subquery, OuterRef
from django.conf import settings
from functools import wraps
import csv
import io
import os
import sys
import subprocess
import threading
import time


def patron_required(view_func):
    """Yetki: yalnızca Patron (superuser) erişebilir; Çalışan dashboard'a yönlenir."""
    @wraps(view_func)
    @login_required
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_superuser:
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return _wrapped


class RememberMeLoginView(LoginView):
    """'Oturum açık kalsın' işaretliyse oturum bu bilgisayarda 1 yıl kalıcı olur
    (uygulama kapanıp açılsa da). İşaretsizse uygulama kapanınca oturum biter."""
    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.POST.get('remember'):
            self.request.session.set_expiry(60 * 60 * 24 * 365)  # 1 yıl
        else:
            self.request.session.set_expiry(0)  # uygulama/tarayıcı kapanınca
        return response


def register_view(request):
    # Two flows share this view:
    if request.user.is_authenticated:
        # 1) Admin "Add User": an authenticated Patron creates an account
        #    WITHOUT switching the current session. Çalışan cannot create users.
        if not request.user.is_superuser:
            return redirect('dashboard')
        if request.method == 'POST':
            form = UserCreationForm(request.POST)
            if form.is_valid():
                form.save()
                return redirect('dashboard')
        else:
            form = UserCreationForm()
        return render(request, 'registration/register.html', {'form': form})

    # 2) Public self-registration from the login screen: a visitor opens their
    #    own account. New self-registered accounts are Çalışan (non-superuser,
    #    non-staff) by default, so registration never grants admin powers.
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('dashboard')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})


@login_required
def dashboard(request):
    products = Product.objects.all()
    total_products = products.count()
    total_stock = sum(p.stock_quantity for p in products)
    low_stock_products = [p for p in products if p.is_low_stock()]
    context = {
        'products': products,
        'total_products': total_products,
        'total_stock': total_stock,
        'low_stock_count': len(low_stock_products),
    }
    return render(request, 'inventory/dashboard.html', context)


@login_required
def product_list(request):
    products = Product.objects.all()

    q = request.GET.get('q', '').strip()
    category = request.GET.get('category', '')
    supplier = request.GET.get('supplier', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    low_stock = request.GET.get('low_stock', '')
    sort = request.GET.get('sort', 'name')

    if q:
        products = products.filter(name__icontains=q)
    if category:
        products = products.filter(category=category)
    if supplier:
        products = products.filter(supplier=supplier)
    if min_price:
        products = products.filter(price__gte=float(min_price))
    if max_price:
        products = products.filter(price__lte=float(max_price))

    # Apply sorting before low_stock (which converts queryset to list)
    if sort == 'stock_asc':
        products = products.order_by('stock_quantity', 'name')
    elif sort == 'stock_desc':
        products = products.order_by('-stock_quantity', 'name')
    elif sort == 'category':
        products = products.order_by('category', 'name')
    elif sort == 'last_movement':
        latest_movement = StockMovement.objects.filter(
            product=OuterRef('pk')
        ).order_by('-date').values('date')[:1]
        products = products.annotate(last_move=Subquery(latest_movement)).order_by('-last_move', 'name')
    else:  # default: name
        products = products.order_by('name')

    if low_stock:
        products = [p for p in products if p.is_low_stock()]

    all_products = Product.objects.all()
    categories = all_products.values_list('category', flat=True).distinct().order_by('category')
    suppliers = all_products.values_list('supplier', flat=True).distinct().order_by('supplier')

    active_filters = {}
    if q: active_filters['Arama'] = q
    if category: active_filters['Kategori'] = category
    if supplier: active_filters['Tedarikçi'] = supplier
    if min_price: active_filters['Min Fiyat'] = f"{min_price} ₺"
    if max_price: active_filters['Max Fiyat'] = f"{max_price} ₺"
    if low_stock: active_filters['Stok'] = 'Düşük Stok'

    # Build query string without 'sort' param (used in template sort links)
    get_params = request.GET.copy()
    get_params.pop('sort', None)
    sort_params = get_params.urlencode()

    context = {
        'products': products,
        'categories': categories,
        'suppliers': suppliers,
        'active_filters': active_filters,
        'q': q,
        'selected_category': category,
        'selected_supplier': supplier,
        'min_price': min_price,
        'max_price': max_price,
        'low_stock': low_stock,
        'current_sort': sort,
        'sort_params': sort_params,
    }
    return render(request, 'inventory/product_list.html', context)
    

@login_required
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save()
            # Başlangıç stoğu varsa stok hareketi kaydet (Stok Hareketleri'nde görünsün)
            if product.stock_quantity and product.stock_quantity > 0:
                StockMovement.objects.create(
                    product=product, change_amount=product.stock_quantity,
                    old_stock=0, new_stock=product.stock_quantity,
                    note='Yeni Ürün (Başlangıç Stoğu)',
                )
            return redirect('product_list')
    else:
        form = ProductForm()
    return render(request, 'inventory/product_form.html', {'form': form})


@login_required
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        old_stock = product.stock_quantity
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            product = form.save()
            # Düzenleme sırasında stok değiştiyse stok hareketi kaydet
            if product.stock_quantity != old_stock:
                StockMovement.objects.create(
                    product=product, change_amount=product.stock_quantity - old_stock,
                    old_stock=old_stock, new_stock=product.stock_quantity,
                    note='Ürün Düzenleme (Stok Değişikliği)',
                )
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'inventory/product_form.html', {'form': form, 'product': product})


@login_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('product_list')
    return render(request, 'inventory/product_confirm_delete.html', {'product': product})


@login_required
def product_increase_stock(request, pk):
    product = get_object_or_404(Product, pk=pk)
    try:
        amount = int(request.POST.get('amount', 1))
        if amount < 1:
            amount = 1
    except (ValueError, TypeError):
        amount = 1
    old = product.stock_quantity
    product.stock_quantity += amount
    product.save()
    StockMovement.objects.create(
        product=product, change_amount=amount,
        old_stock=old, new_stock=product.stock_quantity, note="Stock Increase"
    )
    return HttpResponseRedirect(reverse('product_list'))


@login_required
def product_decrease_stock(request, pk):
    product = get_object_or_404(Product, pk=pk)
    try:
        amount = int(request.POST.get('amount', 1))
        if amount < 1:
            amount = 1
    except (ValueError, TypeError):
        amount = 1
    if product.stock_quantity >= amount:
        old = product.stock_quantity
        product.stock_quantity -= amount
        product.save()
        StockMovement.objects.create(
            product=product, change_amount=-amount,
            old_stock=old, new_stock=product.stock_quantity, note="Stock Decrease"
        )
    return HttpResponseRedirect(reverse('product_list'))


@login_required
def stock_history(request, pk):
    product = get_object_or_404(Product, pk=pk)
    movements = product.movements.order_by('-date')
    return render(request, 'inventory/stock_history.html', {'product': product, 'movements': movements})


@login_required
def all_stock_movements(request):
    movements = StockMovement.objects.select_related('product').order_by('-date')
    return render(request, 'inventory/all_stock_movements.html', {'movements': movements})


# ─── Customer Views ────────────────────────────────────────────────────────────

@patron_required
def customer_list(request):
    q = request.GET.get('q', '').strip()
    sort = request.GET.get('sort', 'name')
    customers = Customer.objects.all()

    if q:
        customers = customers.filter(name__icontains=q)

    # ORM ile sıralanabilenler
    if sort == 'name':
        customers = customers.order_by('name')
    elif sort == 'shop_name':
        customers = customers.order_by('shop_name', 'name')
    elif sort == 'debt_asc':
        customers = customers.order_by('debt', 'name')
    elif sort == 'debt_desc':
        customers = customers.order_by('-debt', 'name')
    elif sort == 'created':
        customers = customers.order_by('-created_at')
    else:
        customers = customers.order_by('name')

    # debt_rating property'e göre sıralama (Python tarafında)
    if sort == 'rating':
        rating_order = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'F': 4}
        customers = sorted(customers, key=lambda c: (rating_order.get(c.debt_rating, 9), c.name))

    get_params = request.GET.copy()
    get_params.pop('sort', None)
    sort_params = get_params.urlencode()

    return render(request, 'inventory/customer_list.html', {
        'customers': customers,
        'q': q,
        'current_sort': sort,
        'sort_params': sort_params,
    })


@patron_required
def customer_create(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('customer_list')
    else:
        form = CustomerForm()
    return render(request, 'inventory/customer_form.html', {'form': form})


@patron_required
def customer_update(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            return redirect('customer_list')
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'inventory/customer_form.html', {'form': form, 'customer': customer})


@patron_required
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.delete()
        return redirect('customer_list')
    return render(request, 'inventory/customer_confirm_delete.html', {'customer': customer})


# ─── Sale Views ────────────────────────────────────────────────────────────────

@login_required
def sale_list(request):
    sales = Sale.objects.select_related('customer').prefetch_related('items').all()
    return render(request, 'inventory/sale_list.html', {'sales': sales})


@login_required
def sale_create(request):
    products = Product.objects.filter(stock_quantity__gt=0).order_by('name')
    customers = Customer.objects.order_by('name')

    if request.method == 'POST':
        customer_id = request.POST.get('customer')
        note = request.POST.get('note', '')
        add_to_debt = request.POST.get('add_to_debt') == 'on'

        # Ürün satırlarını topla
        items = []
        i = 0
        while f'product_{i}' in request.POST:
            product_id = request.POST.get(f'product_{i}')
            quantity = request.POST.get(f'quantity_{i}')
            price = request.POST.get(f'price_{i}')
            if product_id and quantity and price:
                try:
                    items.append({
                        'product': Product.objects.get(pk=product_id),
                        'quantity': int(quantity),
                        'unit_price': float(price),
                    })
                except (Product.DoesNotExist, ValueError):
                    pass
            i += 1

        if not items:
            return render(request, 'inventory/sale_create.html', {
                'products': products,
                'customers': customers,
                'error': 'En az bir ürün eklemelisiniz.',
            })

        # Stok yeterliliği kontrolü
        for item in items:
            if item['product'].stock_quantity < item['quantity']:
                return render(request, 'inventory/sale_create.html', {
                    'products': products,
                    'customers': customers,
                    'error': f'"{item["product"].name}" için yeterli stok yok. Mevcut: {item["product"].stock_quantity}',
                })

        # Satışı oluştur
        customer = Customer.objects.get(pk=customer_id) if customer_id else None
        total = sum(it['quantity'] * it['unit_price'] for it in items)

        sale = Sale.objects.create(
            customer=customer,
            total_amount=round(total, 2),
            add_to_debt=add_to_debt,
            note=note,
        )

        for item in items:
            SaleItem.objects.create(
                sale=sale,
                product=item['product'],
                quantity=item['quantity'],
                unit_price=item['unit_price'],
            )
            # Stok düş
            product = item['product']
            old_stock = product.stock_quantity
            product.stock_quantity -= item['quantity']
            product.save()
            StockMovement.objects.create(
                product=product,
                change_amount=-item['quantity'],
                old_stock=old_stock,
                new_stock=product.stock_quantity,
                note=f'Satış #{sale.pk}',
            )

        # Müşteri borcuna ekle
        if add_to_debt and customer:
            customer.debt = round(customer.debt + total, 2)
            customer.save()

        return redirect('sale_detail', pk=sale.pk)

    return render(request, 'inventory/sale_create.html', {
        'products': products,
        'customers': customers,
    })


@login_required
def sale_detail(request, pk):
    sale = get_object_or_404(Sale.objects.select_related('customer').prefetch_related('items__product'), pk=pk)
    return render(request, 'inventory/sale_detail.html', {'sale': sale})


# ─── Product Import Views ──────────────────────────────────────────────────────

# Tedarikçi "SİPARİŞ FORMU" PDF düzeni — sütunlar x-koordinatına göre ayrılır.
_PDF_UNITS = {
    'adet', 'kutu', 'paket', 'düzine', 'koli', 'gross', 'kg', 'top', 'rulo',
    'set', 'çift', 'takım', 'blok', 'metre', 'mt', 'kavanoz', 'poşet', 'düzine.',
}


def _tr_num(s):
    """'1.200,00' -> '1200.00', '25,00' -> '25.00' (Türkçe sayı normalize)."""
    s = str(s).strip()
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    return s


# PINAR sipariş formu sütun aralıkları (x0)
_PDF_FALLBACK = {'barcode': (35, 98), 'name': (98, 388), 'qty': (388, 420),
                 'unit': (420, 446), 'price': (446, 490)}


def _cluster_rows(words, tol=4):
    words = sorted(words, key=lambda w: (w['top'], w['x0']))
    clusters, cur, ref = [], [], None
    for w in words:
        if ref is None or abs(w['top'] - ref) <= tol:
            cur.append(w)
            ref = w['top'] if ref is None else ref
        else:
            clusters.append(cur)
            cur, ref = [w], w['top']
    if cur:
        clusters.append(cur)
    return clusters


def parse_supplier_pdf(file_obj):
    """Tedarikçi sipariş formu PDF'ini ayrıştırır (PINAR 'SİPARİŞ FORMU' düzeni;
    sütunlar x-koordinatına göre). Taranmış (görüntü) PDF'lerde metin katmanı
    olmadığından boş döner — OCR gerekir."""
    import pdfplumber
    rows = []
    with pdfplumber.open(file_obj) as pdf:
        for pg in pdf.pages:
            words = pg.extract_words(use_text_flow=False)
            if not words:
                continue
            for ws in _cluster_rows(words):
                col = {'barcode': [], 'name': [], 'qty': [], 'unit': [], 'price': []}
                for w in sorted(ws, key=lambda x: x['x0']):
                    x = w['x0']
                    for c, (lo, hi) in _PDF_FALLBACK.items():
                        if lo <= x < hi:
                            col[c].append(w['text'])
                            break
                name = ' '.join(col['name']).strip()
                unit = ' '.join(col['unit']).strip()
                qty = ' '.join(col['qty']).strip()
                price = ' '.join(col['price']).strip()
                barcode = ''.join(col['barcode']).strip()
                if not name or not unit or not price:
                    continue
                if unit.lower() not in _PDF_UNITS:
                    continue
                try:
                    qty_val = str(int(float(_tr_num(qty))))
                except (ValueError, TypeError):
                    qty_val = '0'
                rows.append({
                    'name': name, 'barcode': barcode, 'stock_quantity': qty_val,
                    'unit': unit, 'price': _tr_num(price),
                })
    return rows


def _imp_get(row, *keys, default=''):
    for k in keys:
        if k in row and row[k] is not None and str(row[k]).strip() != '':
            return str(row[k]).strip()
    return default


@login_required
def product_import(request):
    results = None
    preview = None
    errors = []

    # ── Aşama B: düzenlenmiş önizleme onaylandı → ürünleri oluştur/güncelle ──
    if request.method == 'POST' and request.POST.get('confirm') == '1':
        created, updated, skipped = [], [], []
        i = 0
        while f'name_{i}' in request.POST:
            if request.POST.get(f'include_{i}') != 'on':
                i += 1
                continue
            name = request.POST.get(f'name_{i}', '').strip()
            barcode = request.POST.get(f'barcode_{i}', '').strip()
            unit = request.POST.get(f'unit_{i}', 'adet').strip() or 'adet'
            try:
                qty = int(float(request.POST.get(f'qty_{i}', '0').replace(',', '.')))
            except (ValueError, TypeError):
                qty = 0
            try:
                sale_price = float(request.POST.get(f'price_{i}', '0').replace(',', '.'))
            except (ValueError, TypeError):
                sale_price = 0.0

            if not name:
                i += 1
                continue
            if qty <= 0:
                skipped.append({'reason': 'Miktar 0', 'row': name})
                i += 1
                continue

            product = None
            if barcode:
                product = Product.objects.filter(barcode=barcode).first()
            if not product:
                product = Product.objects.filter(name__iexact=name).first()

            if product:
                old_stock = product.stock_quantity
                product.stock_quantity += qty
                if sale_price > 0:
                    product.price = sale_price  # kullanıcının belirlediği satış fiyatı
                product.save()
                StockMovement.objects.create(
                    product=product, change_amount=qty, old_stock=old_stock,
                    new_stock=product.stock_quantity, note='Dosya İçe Aktarma',
                )
                updated.append({'name': product.name, 'added': qty, 'new_stock': product.stock_quantity})
            else:
                try:
                    product = Product.objects.create(
                        name=name, category=request.POST.get(f'category_{i}', '').strip(),
                        price=sale_price, manufacturer=request.POST.get(f'manufacturer_{i}', '').strip(),
                        stock_quantity=qty, unit=unit, barcode=barcode,
                        tax_rate=float(request.POST.get(f'tax_{i}', '20') or 20),
                        supplier=request.POST.get(f'supplier_{i}', '').strip(),
                        shelf_location='', min_stock=0,
                    )
                    StockMovement.objects.create(
                        product=product, change_amount=qty, old_stock=0,
                        new_stock=qty, note='Dosya İçe Aktarma (Yeni Ürün)',
                    )
                    created.append({'name': product.name, 'stock': qty})
                except Exception as e:
                    errors.append(f'"{name}" oluşturulamadı: {str(e)}')
            i += 1

        results = {'created': created, 'updated': updated, 'skipped': skipped, 'errors': errors}
        return render(request, 'inventory/product_import.html', {'results': results})

    # ── Aşama A: dosya yüklendi → ayrıştır, düzenlenebilir önizleme göster (DB'ye yazma) ──
    if request.method == 'POST' and request.FILES.get('import_file'):
        import_file = request.FILES['import_file']
        filename = import_file.name.lower()
        rows = []
        try:
            if filename.endswith('.csv'):
                decoded = import_file.read().decode('utf-8-sig')
                rows = [dict(r) for r in csv.DictReader(io.StringIO(decoded))]
            elif filename.endswith('.xlsx'):
                import openpyxl
                wb = openpyxl.load_workbook(import_file, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        rows.append(dict(zip(headers, row)))
            elif filename.endswith('.pdf'):
                rows = parse_supplier_pdf(import_file)
                if not rows:
                    errors.append('PDF okundu ama ürün satırı bulunamadı. Desteklenen düzen: tedarikçi sipariş formu.')
            else:
                errors.append('Desteklenmeyen dosya formatı. Lütfen .csv, .xlsx veya .pdf dosyası yükleyin.')
        except Exception as e:
            errors.append(f'Dosya okunurken hata oluştu: {str(e)}')

        # Başlık satırı / teknik-ad satırı sayılacak değerler (atlanır)
        _header_like = {
            'name', 'ürün adı', 'urun adi', 'stok adı', 'stok adi', 'stok_adi',
            'stok adı', 'ürün adi', 'stokadı',
        }
        preview = []
        for row in rows:
            # Ad: 'Ürün Adı' veya 'Stok Adı' (farklı tedarikçi/program başlıkları)
            name = _imp_get(row, 'name', 'Ürün Adı', 'Stok Adı', 'STOK_ADI', 'STOK ADI')
            if not name or name.strip().lower() in _header_like:
                continue
            try:
                qty = int(float(_imp_get(row, 'stock_quantity', 'Stok Miktarı', 'Miktar', 'MIKTAR', default='0').replace(',', '.')))
            except (ValueError, TypeError):
                qty = 0
            try:
                purchase = float(_imp_get(row, 'price', 'Fiyat', 'Birim Fiyat', 'BIRIM_FIYAT', 'Birim Fiyatı', default='0').replace(',', '.'))
            except (ValueError, TypeError):
                purchase = 0.0
            barcode = _imp_get(row, 'barcode', 'Barkod', 'BARKOD')
            existing = None
            if barcode:
                existing = Product.objects.filter(barcode=barcode).first()
            if not existing:
                existing = Product.objects.filter(name__iexact=name).first()
            preview.append({
                'name': name, 'barcode': barcode,
                'unit': _imp_get(row, 'unit', 'Birim', 'BIRIM', default='adet'),
                'category': _imp_get(row, 'category', 'Kategori'),
                'manufacturer': _imp_get(row, 'manufacturer', 'Üretici'),
                'supplier': _imp_get(row, 'supplier', 'Tedarikçi'),
                'tax_rate': _imp_get(row, 'tax_rate', 'KDV Oranı', default='20'),
                'qty': qty,
                'purchase_price': purchase,
                'exists': bool(existing),
                'current_stock': existing.stock_quantity if existing else 0,
            })

    return render(request, 'inventory/product_import.html', {'results': results, 'preview': preview, 'preview_errors': errors})


@login_required
def product_intake(request):
    """Hızlı ürün girişi / alım sayfası: yeni gelen ürünleri liste halinde elle gir.
    Sadece barkod, ad, miktar, alım fiyatı (referans), satış fiyatı. Kaydedince
    eşleşen ürünün stoğu artar (+satış fiyatı güncellenir), yenisi oluşturulur."""
    saved = None
    if request.method == 'POST':
        added, updated, skipped = [], [], []
        i = 0
        while f'name_{i}' in request.POST:
            name = request.POST.get(f'name_{i}', '').strip()
            if not name:
                i += 1
                continue
            barcode = request.POST.get(f'barcode_{i}', '').strip()
            try:
                qty = int(float(request.POST.get(f'qty_{i}', '1').replace(',', '.')))
            except (ValueError, TypeError):
                qty = 1
            if qty < 1:
                qty = 1
            try:
                sell = float(request.POST.get(f'sell_{i}', '0').replace(',', '.'))
            except (ValueError, TypeError):
                sell = 0.0

            product = None
            if barcode:
                product = Product.objects.filter(barcode=barcode).first()
            if not product:
                product = Product.objects.filter(name__iexact=name).first()

            if product:
                old = product.stock_quantity
                product.stock_quantity += qty
                if sell > 0:
                    product.price = sell
                product.save()
                StockMovement.objects.create(
                    product=product, change_amount=qty, old_stock=old,
                    new_stock=product.stock_quantity, note='Alım',
                )
                updated.append({'name': product.name, 'added': qty, 'new_stock': product.stock_quantity})
            else:
                product = Product.objects.create(
                    name=name, barcode=barcode, price=sell, stock_quantity=qty,
                    unit='adet', category='', manufacturer='', tax_rate=20.0,
                    supplier='', shelf_location='', min_stock=0,
                )
                StockMovement.objects.create(
                    product=product, change_amount=qty, old_stock=0,
                    new_stock=qty, note='Alım (Yeni Ürün)',
                )
                added.append({'name': product.name, 'stock': qty})
            i += 1

        saved = {'added': added, 'updated': updated, 'skipped': skipped}

    return render(request, 'inventory/product_intake.html', {'saved': saved})


@login_required
def product_import_template(request):
    """Boş CSV şablon dosyası indir."""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="urun_import_sablonu.csv"'
    writer = csv.writer(response)
    writer.writerow(['name', 'category', 'price', 'manufacturer', 'stock_quantity', 'unit', 'barcode', 'tax_rate', 'supplier', 'shelf_location', 'min_stock'])
    writer.writerow(['Örnek Ürün A4 Kağıt', 'Kağıt', '25.50', 'Papirus', '100', 'paket', '8680000000001', '20', 'Örnek Tedarikçi', 'A-01', '10'])
    writer.writerow(['Tükenmez Kalem', 'Kalem', '3.75', 'Bic', '500', 'adet', '8680000000002', '20', '', 'B-03', '50'])
    return response


# ─── Barcode Memory (Reference) Views ──────────────────────────────────────────

@login_required
def product_search(request):
    """Canlı arama (otomatik tamamlama): isim veya barkoda göre ürün önerir.
    scope=sale -> sadece stokta olanlar; scope=intake -> ürünler + barkod hafızası."""
    from django.db.models import Q
    q = request.GET.get('q', '').strip()
    scope = request.GET.get('scope', 'sale')
    results = []
    if len(q) < 2:
        return JsonResponse({'results': results})

    qs = Product.objects.filter(Q(name__icontains=q) | Q(barcode__icontains=q))
    if scope == 'sale':
        qs = qs.filter(stock_quantity__gt=0)
    for p in qs.order_by('name')[:10]:
        results.append({
            'type': 'product', 'id': p.pk, 'name': p.name, 'barcode': p.barcode,
            'price': p.price, 'stock': p.stock_quantity, 'unit': p.unit,
        })

    if scope == 'intake':
        seen = {r['barcode'] for r in results if r['barcode']}
        refs = ReferenceBarcode.objects.filter(Q(name__icontains=q) | Q(barcode__icontains=q)).order_by('name')[:10]
        for r in refs:
            if r.barcode and r.barcode in seen:
                continue
            results.append({'type': 'reference', 'name': r.name, 'barcode': r.barcode, 'unit': r.unit})

    return JsonResponse({'results': results[:12]})


@login_required
def barcode_lookup(request, code):
    """Yerel barkod hafızasından arama. Sadece tek ürün döner; dışarıya istek gitmez."""
    code = (code or '').strip()
    data = {'found': False, 'already_in_stock': False}

    existing = Product.objects.filter(barcode=code).first()
    if existing:
        data['already_in_stock'] = True
        data['existing_name'] = existing.name

    ref = ReferenceBarcode.objects.filter(barcode=code).first()
    if ref:
        data.update({
            'found': True,
            'name': ref.name,
            'manufacturer': ref.manufacturer,
            'category': ref.category,
            'unit': ref.unit,
            'tax_rate': ref.tax_rate,
        })
    return JsonResponse(data)


@patron_required
def reference_import(request):
    """Referans barkod hafızasına CSV/XLSX ile toplu kayıt ekle."""
    results = None
    if request.method == 'POST' and request.FILES.get('import_file'):
        import_file = request.FILES['import_file']
        filename = import_file.name.lower()
        rows = []
        errors = []
        try:
            if filename.endswith('.csv'):
                decoded = import_file.read().decode('utf-8-sig')
                rows = [dict(r) for r in csv.DictReader(io.StringIO(decoded))]
            elif filename.endswith('.xlsx'):
                import openpyxl
                wb = openpyxl.load_workbook(import_file, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        rows.append(dict(zip(headers, row)))
            else:
                errors.append('Desteklenmeyen dosya formatı. .csv veya .xlsx yükleyin.')
        except Exception as e:
            errors.append(f'Dosya okunamadı: {str(e)}')

        def gv(row, *keys, default=''):
            for k in keys:
                if k in row and row[k] is not None and str(row[k]).strip() != '':
                    return str(row[k]).strip()
            return default

        added, updated, skipped = 0, 0, 0
        for row in rows:
            code = gv(row, 'barcode', 'Barkod')
            name = gv(row, 'name', 'Ürün Adı')
            if not code or not name or code.lower() in ('barcode', 'barkod'):
                skipped += 1
                continue
            try:
                tax = float(gv(row, 'tax_rate', 'KDV Oranı', default='20').replace(',', '.'))
            except (ValueError, TypeError):
                tax = 20.0
            obj, created = ReferenceBarcode.objects.update_or_create(
                barcode=code,
                defaults={
                    'name': name,
                    'manufacturer': gv(row, 'manufacturer', 'Üretici'),
                    'category': gv(row, 'category', 'Kategori'),
                    'unit': gv(row, 'unit', 'Birim', default='adet'),
                    'tax_rate': tax,
                },
            )
            if created:
                added += 1
            else:
                updated += 1

        results = {'added': added, 'updated': updated, 'skipped': skipped,
                   'errors': errors, 'total': ReferenceBarcode.objects.count()}

    return render(request, 'inventory/reference_import.html', {'results': results})


@patron_required
def reference_import_template(request):
    """Referans barkod için boş CSV şablonu."""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="referans_barkod_sablonu.csv"'
    writer = csv.writer(response)
    writer.writerow(['barcode', 'name', 'manufacturer', 'category', 'unit', 'tax_rate'])
    writer.writerow(['8680000000001', 'Örnek A4 Kağıt', 'Papirus', 'Kağıt', 'paket', '20'])
    writer.writerow(['8680000000002', 'Örnek Tükenmez Kalem', 'Bic', 'Kalem', 'adet', '20'])
    return response


# ─── Uygulama Güncelleme (uygulama içinden) ─────────────────────────────────────

@patron_required
def update_app(request):
    """Uygulama içinden GitHub'dan güncelleme: git pull + paketler + migrate + collectstatic.
    Sadece Patron. Kod değişiklikleri için sonrasında yeniden başlatma gerekir."""
    result = None
    if request.method == 'POST':
        repo_root = os.path.dirname(settings.BASE_DIR)  # .../Papirus-sys-main (.git burada)
        base = str(settings.BASE_DIR)
        py = sys.executable
        steps = []

        def run(label, args, cwd):
            try:
                p = subprocess.run(args, cwd=cwd, capture_output=True, text=True,
                                   timeout=300, encoding='utf-8', errors='replace')
                out = ((p.stdout or '') + (p.stderr or '')).strip()
                steps.append({'label': label, 'ok': p.returncode == 0, 'out': out or '(çıktı yok)'})
                return p.returncode == 0
            except FileNotFoundError:
                steps.append({'label': label, 'ok': False, 'out': 'Komut bulunamadı (git kurulu mu?).'})
                return False
            except Exception as e:
                steps.append({'label': label, 'ok': False, 'out': str(e)})
                return False

        ok = run('Yeni sürüm indiriliyor (git pull)', ['git', 'pull'], repo_root)
        updated = ok and 'Already up to date' not in steps[-1]['out'] and 'güncel' not in steps[-1]['out'].lower()
        if ok and updated:
            run('Paketler güncelleniyor', [py, '-m', 'pip', 'install', '-r', os.path.join(base, 'requirements.txt'), '--quiet'], base)
            run('Veritabanı güncelleniyor (migrate)', [py, os.path.join(base, 'manage.py'), 'migrate', '--noinput'], base)
            run('Statik dosyalar', [py, os.path.join(base, 'manage.py'), 'collectstatic', '--noinput'], base)
        result = {'steps': steps, 'ok': ok, 'updated': updated}
    return render(request, 'inventory/update_app.html', {'result': result})


@patron_required
def restart_app(request):
    """Uygulamayı yeniden başlatır: yeni süreç başlatıp mevcut süreci kapatır."""
    if request.method == 'POST':
        def _restart():
            time.sleep(1.0)
            try:
                flags = getattr(subprocess, 'DETACHED_PROCESS', 0) | getattr(subprocess, 'CREATE_NEW_PROCESS_GROUP', 0)
                subprocess.Popen([sys.executable, os.path.join(str(settings.BASE_DIR), 'launcher.py')],
                                 cwd=str(settings.BASE_DIR), close_fds=True, creationflags=flags)
            except Exception:
                pass
            os._exit(0)
        threading.Thread(target=_restart, daemon=True).start()
        return render(request, 'inventory/restarting.html', {})
    return redirect('update_app')
